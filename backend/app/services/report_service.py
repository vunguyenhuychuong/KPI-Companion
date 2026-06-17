"""Xuat bao cao Excel theo mau danh gia: KPI, tien do, bang chung, viec phat sinh, lich su."""
import io
import json
import re
from collections import defaultdict
from datetime import date

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from sqlalchemy import select
from sqlalchemy.orm import Session

from .. import models, schemas
from ..agent.llm import call_json
from . import kpi_service

HEADER_FILL = PatternFill("solid", fgColor="1F4E79")
HEADER_FONT = Font(bold=True, color="FFFFFF", size=11)
TITLE_FONT = Font(bold=True, size=14, color="1F4E79")
THIN = Side(style="thin", color="B0B0B0")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
HEALTH_FILLS = {
    "green": PatternFill("solid", fgColor="C6EFCE"),
    "yellow": PatternFill("solid", fgColor="FFEB9C"),
    "red": PatternFill("solid", fgColor="FFC7CE"),
}


def _style_header(ws, row: int, n_cols: int):
    for c in range(1, n_cols + 1):
        cell = ws.cell(row=row, column=c)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.border = BORDER
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)


APPRAISAL_BLUE = PatternFill("solid", fgColor="B8CCE4")
RED_BOLD = Font(bold=True, color="FF0000", size=10)


APPRAISAL_SELF_REVIEW_SYSTEM = """
Bạn viết nội dung cho cột "Nhân viên tự nhận xét/Employee self-assessment*" trong file Performance Appraisal.

Quy tắc bắt buộc:
- Chỉ dùng dữ liệu được cung cấp trong `verified_result` và `confirmed_evidence`.
- `context` chỉ giúp hiểu KPI, KHÔNG phải bằng chứng hoàn thành. Không lấy chi tiết từ context để biến thành thành tích.
- Chi tiết cụ thể như tên chứng chỉ, công nghệ, dự án, loại đầu việc, nguồn việc chỉ được nêu nếu xuất hiện trong `confirmed_evidence`.
- Nếu không có `confirmed_evidence`, không tự thêm ví dụ, tên công nghệ, tên chứng chỉ hoặc dự án.
- Nếu chỉ có số tổng, viết rằng hệ thống mới ghi nhận ở mức tổng; không phủ nhận kết quả nhưng cũng không thêm chi tiết.
- Viết tự nhiên như nhân viên tự nhận xét, không dùng bullet, không markdown, không dấu *.
- Mỗi KPI tối đa 2 câu, tối đa 65 từ. Tránh lặp câu hoặc lặp ý.
- Không nêu ID nội bộ.

Trả về JSON đúng dạng:
{"comments":[{"index":1,"comment":"..."}]}
""".strip()


APPRAISAL_OVERVIEW_SYSTEM = """
Bạn viết 2 ô đầu file Performance Appraisal:
1. "Nhân viên nhận xét tổng quan/Overall Employee Comment"
2. "Nhu cầu phát triển của nhân viên/Development Needs"

Quy tắc bắt buộc:
- Chỉ dùng dữ liệu được cung cấp trong summary và confirmed evidence. Không bịa thành tích, số liệu, nguồn, dự án, cảm xúc hoặc bối cảnh.
- Context định nghĩa KPI không phải bằng chứng hoàn thành; không biến context thành thành tích đã làm.
- Được suy luận ngắn gọn từ khoảng trống/rủi ro có thật, ví dụ thiếu bằng chứng, tiến độ thấp, KPI chưa có ghi nhận kỳ.
- Viết ở ngôi thứ nhất, tự nhiên như nhân viên tự đánh giá, không markdown, không bullet, không dấu *.
- Mỗi ô 1-2 câu, tối đa 70 từ.
- Không nêu ID nội bộ.

Trả về JSON đúng dạng:
{"overall_comment":"...","development_needs":"..."}
""".strip()


SOURCE_LABELS = {
    "chat": "nhật ký công việc",
    "csv": "file import",
    "gmail": "Gmail",
    "calendar": "Calendar",
    "sheets": "Google Sheets",
    "notion": "Notion",
    "slack": "Slack",
    "outlook": "Outlook",
    "agent_loop": "Agent tự chủ",
    "manual": "ghi nhận thủ công",
}


GROUNDING_GENERIC_WORDS = {
    "báo",
    "bằng",
    "bổ",
    "bộ",
    "các",
    "cáo",
    "cần",
    "cập",
    "chỉ",
    "chi",
    "cho",
    "chưa",
    "chung",
    "chứng",
    "có",
    "cơ",
    "còn",
    "của",
    "cụ",
    "đang",
    "đạt",
    "đã",
    "để",
    "đến",
    "đều",
    "định",
    "đo",
    "do",
    "dõi",
    "dữ",
    "đủ",
    "được",
    "dựa",
    "giá",
    "ghi",
    "hơn",
    "hiện",
    "hoàn",
    "hoặc",
    "khi",
    "không",
    "kỳ",
    "kpi",
    "là",
    "lại",
    "làm",
    "lên",
    "liệu",
    "mới",
    "mục",
    "mức",
    "này",
    "nên",
    "nếu",
    "nguồn",
    "nhận",
    "nhật",
    "nhưng",
    "nội",
    "phản",
    "sâu",
    "số",
    "so",
    "sở",
    "sung",
    "target",
    "theo",
    "thể",
    "thêm",
    "thực",
    "tiêu",
    "tiết",
    "tiếp",
    "tiến",
    "tôi",
    "trạng",
    "trên",
    "trong",
    "từ",
    "tự",
    "tổng",
    "và",
    "vào",
    "với",
    "vượt",
    "xác",
}


def _plain(text: str | None, limit: int = 240) -> str:
    value = re.sub(r"[*_`#>\[\]]+", "", str(text or ""))
    value = re.sub(r"\s+", " ", value).strip()
    value = re.sub(r"([.!?])(?=\S)", r"\1 ", value)
    if len(value) <= limit:
        return value
    return value[: limit - 1].rstrip() + "…"


def _clean_comment(text: str | None, limit: int = 520) -> str:
    value = _plain(text, limit * 2)
    if not value:
        return ""
    parts = re.split(r"(?<=[.!?])\s+", value)
    cleaned: list[str] = []
    seen: set[str] = set()
    for part in parts:
        sentence = part.strip()
        if not sentence:
            continue
        key = re.sub(r"\W+", "", sentence, flags=re.UNICODE).lower()
        if key and key in seen:
            continue
        seen.add(key)
        cleaned.append(sentence)
    return _plain(" ".join(cleaned), limit)


def _word_tokens(text: str | None) -> set[str]:
    return set(re.findall(r"[0-9A-Za-zÀ-ỹĐđ]+", str(text or "").lower()))


def _has_confirmed_evidence(
    work_items: list[models.WorkItem],
    period_metrics: list[models.KPIPeriodMetric],
) -> bool:
    return bool(work_items or period_metrics)


def _evidence_texts(
    work_items: list[models.WorkItem],
    period_metrics: list[models.KPIPeriodMetric],
) -> list[str]:
    texts: list[str] = []
    for item in work_items:
        item_date = item.work_date or item.created_at.date()
        texts.extend(
            [
                item.title,
                item.detail,
                schemas.STATUS_LABELS.get(item.status, item.status),
                _source_label(item.source),
                item_date.isoformat(),
                f"{item.progress_delta:g}",
            ]
        )
    for metric in period_metrics:
        texts.extend(
            [
                metric.period_key,
                metric.period_type,
                _source_label(metric.source_type),
                f"{metric.actual_value:g}",
                f"{metric.target_value:g}",
                f"{metric.attainment_pct:g}",
            ]
        )
    return [str(t) for t in texts if str(t or "").strip()]


def _is_grounded_comment(
    comment: str,
    kpi: models.KPI,
    work_items: list[models.WorkItem],
    period_metrics: list[models.KPIPeriodMetric],
) -> bool:
    if not _has_confirmed_evidence(work_items, period_metrics):
        return False
    allowed_texts = [
        kpi.name,
        kpi.objective_name or "",
        kpi.unit,
        _achievement_sentence(kpi),
        _progress_note(kpi),
        f"{kpi.current_value:g}",
        f"{kpi.target_value:g}",
        f"{kpi.progress:g}",
        *_evidence_texts(work_items, period_metrics),
    ]
    allowed_tokens = set(GROUNDING_GENERIC_WORDS)
    for text in allowed_texts:
        allowed_tokens.update(_word_tokens(text))
    unsupported = {
        token
        for token in _word_tokens(comment)
        if token not in allowed_tokens
    }
    return not unsupported


def _is_grounded_overall_text(
    text: str,
    kpis: list[models.KPI],
    work_by_kpi: dict[int, list[models.WorkItem]],
    metrics_by_kpi: dict[int, list[models.KPIPeriodMetric]],
) -> bool:
    evidence_count = sum(len(work_by_kpi.get(k.id, [])) + len(metrics_by_kpi.get(k.id, [])) for k in kpis)
    if evidence_count == 0:
        return False
    allowed_texts: list[str] = [
        str(len(kpis)),
        str(evidence_count),
        f"{sum(k.progress for k in kpis) / len(kpis):.0f}" if kpis else "0",
        str(sum(1 for k in kpis if k.progress >= 100)),
        str(sum(1 for k in kpis if k.progress < 70)),
    ]
    for kpi in kpis:
        allowed_texts.extend(
            [
                kpi.name,
                kpi.objective_name or "",
                kpi.unit,
                _achievement_sentence(kpi),
                _progress_note(kpi),
                *_evidence_texts(work_by_kpi.get(kpi.id, []), metrics_by_kpi.get(kpi.id, [])),
            ]
        )
    allowed_tokens = set(GROUNDING_GENERIC_WORDS)
    for allowed in allowed_texts:
        allowed_tokens.update(_word_tokens(allowed))
    unsupported = {
        token
        for token in _word_tokens(text)
        if token not in allowed_tokens
    }
    return not unsupported


def _source_label(source: str | None) -> str:
    return SOURCE_LABELS.get((source or "").strip().lower(), source or "nguồn ghi nhận")


def _progress_note(kpi: models.KPI) -> str:
    over = " — vượt chỉ tiêu" if kpi.progress > 100 else ""
    if kpi.unit == "%":
        return f"tiến độ hiện tại {kpi.current_value:g}%{over}"
    return f"thực đạt {kpi.current_value:g}/{kpi.target_value:g} {kpi.unit} ({kpi.progress:.0f}%){over}"


def _achievement_sentence(kpi: models.KPI) -> str:
    if kpi.unit == "%":
        base = f"Đạt {kpi.current_value:g}%"
        if kpi.target_value and kpi.current_value > kpi.target_value:
            return f"{base}, vượt {kpi.current_value - kpi.target_value:g} điểm phần trăm so với mục tiêu"
        return f"{base} so với mục tiêu {kpi.target_value:g}%"
    base = f"Hoàn thành {kpi.current_value:g}/{kpi.target_value:g} {kpi.unit}"
    if kpi.target_value and kpi.current_value > kpi.target_value:
        return f"{base}, vượt {kpi.current_value - kpi.target_value:g} {kpi.unit} so với mục tiêu"
    return f"{base} ({kpi.progress:.0f}%)"


def _group_appraisal_evidence(
    db: Session, user_id: int, kpis: list[models.KPI]
) -> tuple[dict[int, list[models.WorkItem]], dict[int, list[models.KPIPeriodMetric]]]:
    kpi_ids = [k.id for k in kpis]
    if not kpi_ids:
        return {}, {}

    work_by_kpi: dict[int, list[models.WorkItem]] = defaultdict(list)
    work_items = list(
        db.scalars(
            select(models.WorkItem)
            .where(
                models.WorkItem.user_id == user_id,
                models.WorkItem.confirmed == True,  # noqa: E712
                models.WorkItem.kpi_id.in_(kpi_ids),
            )
            .order_by(models.WorkItem.work_date.desc().nullslast(), models.WorkItem.created_at.desc())
        )
    )
    for item in work_items:
        if item.kpi_id and len(work_by_kpi[item.kpi_id]) < 8:
            work_by_kpi[item.kpi_id].append(item)

    metrics_by_kpi: dict[int, list[models.KPIPeriodMetric]] = defaultdict(list)
    period_metrics = list(
        db.scalars(
            select(models.KPIPeriodMetric)
            .where(
                models.KPIPeriodMetric.user_id == user_id,
                models.KPIPeriodMetric.confirmed == True,  # noqa: E712
                models.KPIPeriodMetric.kpi_id.in_(kpi_ids),
            )
            .order_by(models.KPIPeriodMetric.period_start.desc(), models.KPIPeriodMetric.updated_at.desc())
        )
    )
    for metric in period_metrics:
        if len(metrics_by_kpi[metric.kpi_id]) < 5:
            metrics_by_kpi[metric.kpi_id].append(metric)

    return dict(work_by_kpi), dict(metrics_by_kpi)


def _deterministic_self_assessment(
    kpi: models.KPI,
    work_items: list[models.WorkItem],
    period_metrics: list[models.KPIPeriodMetric],
) -> str:
    progress = _achievement_sentence(kpi)
    metric = period_metrics[0] if period_metrics else None
    done_items = [w for w in work_items if w.status == "da_lam"]
    item = done_items[0] if done_items else (work_items[0] if work_items else None)

    if item:
        item_date = item.work_date or item.created_at.date()
        source = _source_label(item.source)
        status = schemas.STATUS_LABELS.get(item.status, item.status)
        return _clean_comment(
            f"{progress}. Bằng chứng gần nhất là \"{item.title}\" "
            f"ngày {item_date.isoformat()} từ {source}, trạng thái {status}.",
            420,
        )
    if metric:
        return _clean_comment(
            f"{progress}. Kỳ {metric.period_key} có số thực tế "
            f"{metric.actual_value:g}/{metric.target_value:g}, đạt {metric.attainment_pct:.0f}% "
            f"từ {_source_label(metric.source_type)}.",
            420,
        )
    return _clean_comment(
        f"{progress}. Hiện dữ liệu đã xác nhận mới có số tổng trên KPI, chưa có bằng chứng công việc chi tiết để diễn giải sâu hơn.",
        420,
    )


def _appraisal_payload_item(
    index: int,
    kpi: models.KPI,
    work_items: list[models.WorkItem],
    period_metrics: list[models.KPIPeriodMetric],
) -> dict:
    evidence = []
    for item in work_items[:5]:
        item_date = item.work_date or item.created_at.date()
        evidence.append(
            {
                "type": "work_item",
                "title": _plain(item.title, 160),
                "detail": _plain(item.detail, 180),
                "status": schemas.STATUS_LABELS.get(item.status, item.status),
                "date": item_date.isoformat(),
                "source": _source_label(item.source),
                "progress_delta": item.progress_delta,
            }
        )
    for metric in period_metrics[:3]:
        evidence.append(
            {
                "type": "period_metric",
                "period": metric.period_key,
                "actual": metric.actual_value,
                "target": metric.target_value,
                "attainment_pct": metric.attainment_pct,
                "source": _source_label(metric.source_type),
            }
        )

    return {
        "index": index,
        "context": {
            "kpi_name": _plain(kpi.name, 180),
            "objective": _plain(kpi.objective_name, 180),
            "target": _plain(kpi.target, 240),
            "description": _plain(kpi.description, 280),
            "note": "Context chỉ là định nghĩa KPI, không phải bằng chứng hoàn thành.",
        },
        "verified_result": {
            "unit": kpi.unit,
            "target_value": kpi.target_value,
            "current_value": kpi.current_value,
            "progress_pct": kpi.progress,
            "achievement_summary": _achievement_sentence(kpi),
            "progress_note": _progress_note(kpi),
        },
        "has_confirmed_evidence": _has_confirmed_evidence(work_items, period_metrics),
        "confirmed_evidence": evidence,
    }


def _ai_self_assessment_comments(db: Session, user_id: int, kpis: list[models.KPI]) -> dict[int, str]:
    work_by_kpi, metrics_by_kpi = _group_appraisal_evidence(db, user_id, kpis)
    fallback = {
        k.id: _deterministic_self_assessment(k, work_by_kpi.get(k.id, []), metrics_by_kpi.get(k.id, []))
        for k in kpis
    }
    index_to_kpi = {i: k for i, k in enumerate(kpis, 1)}
    payload = []
    for i, k in index_to_kpi.items():
        work_items = work_by_kpi.get(k.id, [])
        period_metrics = metrics_by_kpi.get(k.id, [])
        if not _has_confirmed_evidence(work_items, period_metrics):
            continue
        payload.append(_appraisal_payload_item(i, k, work_items, period_metrics))
    if not payload:
        return fallback

    by_index: dict[int, str] = {}
    for start in range(0, len(payload), 10):
        chunk = payload[start : start + 10]
        try:
            result = call_json(
                APPRAISAL_SELF_REVIEW_SYSTEM,
                json.dumps({"today": date.today().isoformat(), "kpis": chunk}, ensure_ascii=False),
                temperature=0.2,
                max_tokens=min(2600, 260 + len(chunk) * 190),
            )
        except Exception:
            continue
        if not isinstance(result, dict) or not isinstance(result.get("comments"), list):
            continue
        for item in result["comments"]:
            if not isinstance(item, dict):
                continue
            try:
                idx = int(item.get("index"))
            except (TypeError, ValueError):
                continue
            comment = _clean_comment(item.get("comment"), 520)
            kpi = index_to_kpi.get(idx)
            if kpi and comment and _is_grounded_comment(
                comment,
                kpi,
                work_by_kpi.get(kpi.id, []),
                metrics_by_kpi.get(kpi.id, []),
            ):
                by_index[idx] = comment

    comments = fallback.copy()
    for i, k in enumerate(kpis, 1):
        if by_index.get(i):
            comments[k.id] = by_index[i]
    return comments


def _fallback_overall_appraisal(
    kpis: list[models.KPI],
    work_by_kpi: dict[int, list[models.WorkItem]],
    metrics_by_kpi: dict[int, list[models.KPIPeriodMetric]],
) -> tuple[str, str]:
    if not kpis:
        return (
            "Tôi chưa có dữ liệu KPI đã ghi nhận trong kỳ nên chưa đủ cơ sở để tự đánh giá tổng quan.",
            "Tôi cần bổ sung KPI, mục tiêu đo lường và bằng chứng công việc đã xác nhận để có cơ sở đánh giá kỳ này.",
        )

    avg_progress = sum(k.progress for k in kpis) / len(kpis)
    completed = sum(1 for k in kpis if k.progress >= 100)
    at_risk = sum(1 for k in kpis if k.progress < 70)
    evidence_count = sum(len(work_by_kpi.get(k.id, [])) + len(metrics_by_kpi.get(k.id, [])) for k in kpis)
    overall = (
        f"Tôi đang theo dõi {len(kpis)} KPI với tiến độ trung bình khoảng {avg_progress:.0f}%; "
        f"{completed} KPI đã đạt hoặc vượt mục tiêu và {at_risk} KPI còn dưới 70%. "
        f"Nội dung này dựa trên {evidence_count} bằng chứng hoặc kỳ ghi nhận đã xác nhận."
    )
    if evidence_count == 0:
        overall = (
            f"Tôi đang theo dõi {len(kpis)} KPI với tiến độ trung bình khoảng {avg_progress:.0f}%, "
            "nhưng chưa có nhật ký công việc hoặc nguồn tích hợp đã xác nhận để đánh giá sâu hơn."
        )

    if at_risk:
        needs = (
            "Tôi cần ưu tiên các KPI còn dưới 70%, cập nhật bằng chứng công việc đều hơn và rà lại cách đo "
            "để tiến độ phản ánh đúng thực tế trong kỳ tiếp theo."
        )
    elif evidence_count < len(kpis):
        needs = (
            "Tôi cần bổ sung bằng chứng định kỳ cho từng KPI để phần tự đánh giá có cơ sở rõ hơn, "
            "không chỉ dựa vào số tiến độ hiện tại."
        )
    else:
        needs = (
            "Tôi cần tiếp tục duy trì nhịp ghi nhận bằng chứng, chuẩn hóa nguồn dữ liệu và theo dõi sớm các KPI "
            "có dấu hiệu chậm tiến độ."
        )
    return _clean_comment(overall, 520), _clean_comment(needs, 520)


def _ai_overall_appraisal(db: Session, user_id: int, kpis: list[models.KPI]) -> tuple[str, str]:
    work_by_kpi, metrics_by_kpi = _group_appraisal_evidence(db, user_id, kpis)
    fallback_overall, fallback_needs = _fallback_overall_appraisal(kpis, work_by_kpi, metrics_by_kpi)
    if not kpis:
        return fallback_overall, fallback_needs

    summary = {
        "kpi_count": len(kpis),
        "average_progress_pct": round(sum(k.progress for k in kpis) / len(kpis), 1),
        "completed_or_over_target": sum(1 for k in kpis if k.progress >= 100),
        "below_70_pct": sum(1 for k in kpis if k.progress < 70),
        "confirmed_evidence_count": sum(
            len(work_by_kpi.get(k.id, [])) + len(metrics_by_kpi.get(k.id, [])) for k in kpis
        ),
    }
    if summary["confirmed_evidence_count"] == 0:
        return fallback_overall, fallback_needs

    payload = [
        _appraisal_payload_item(i, k, work_by_kpi.get(k.id, []), metrics_by_kpi.get(k.id, []))
        for i, k in enumerate(kpis[:20], 1)
        if _has_confirmed_evidence(work_by_kpi.get(k.id, []), metrics_by_kpi.get(k.id, []))
    ]
    try:
        result = call_json(
            APPRAISAL_OVERVIEW_SYSTEM,
            json.dumps({"today": date.today().isoformat(), "summary": summary, "kpis": payload}, ensure_ascii=False),
            temperature=0.2,
            max_tokens=700,
        )
    except Exception:
        return fallback_overall, fallback_needs

    if not isinstance(result, dict):
        return fallback_overall, fallback_needs
    overall = _clean_comment(result.get("overall_comment"), 520) or fallback_overall
    needs = _clean_comment(result.get("development_needs"), 520) or fallback_needs
    if not _is_grounded_overall_text(overall, kpis, work_by_kpi, metrics_by_kpi):
        overall = fallback_overall
    if not _is_grounded_overall_text(needs, kpis, work_by_kpi, metrics_by_kpi):
        needs = fallback_needs
    return overall, needs


def export_appraisal_excel(db: Session, user: models.User, cycle_id: int | None = None) -> bytes:
    """Xuat dung MAU PERFORMANCE APPRAISAL cua cong ty — import nguoc lai duoc.

    Chua co KPI -> xuat template trong (giu nguyen tieu de, quy tac, header)
    de nguoi dung dien tay hoac doi chieu.
    """
    kpis = kpi_service.get_active_kpis(db, user.id, cycle_id=cycle_id)
    objective_query = select(models.Objective).where(
        models.Objective.user_id == user.id,
        models.Objective.archived == False,  # noqa: E712
    )
    if cycle_id is not None:
        objective_query = objective_query.where(models.Objective.cycle_id == cycle_id)
    objectives = list(db.scalars(objective_query))
    year = kpis[0].year if kpis else date.today().year
    overall_comment, development_needs = _ai_overall_appraisal(db, user.id, kpis)
    self_assessment_comments = _ai_self_assessment_comments(db, user.id, kpis) if kpis else {}

    wb = Workbook()
    ws = wb.active
    ws.title = "Performance Appraisal"

    # ===== Phan dau: thong tin nhan vien + quy tac =====
    ws["A1"] = f"{year} Performance Appraisal"
    ws["A1"].font = Font(bold=True, size=18)
    ws["A2"] = "Mã nhân viên/Employee code"
    ws["B2"] = user.name or ""
    ws["A3"] = "Họ tên nhân viên/Employee name"
    ws["B3"] = user.employee_code or ""
    for r in (2, 3):
        ws.cell(row=r, column=1).font = Font(bold=True)
    ws["A4"] = (
        "(*) Trường thông tin bắt buộc/required field\n"
        "- Tổng tỷ trọng KPIs của mỗi Objectives phải bằng 100%/The sum of all KPIs of each Objective is 100%\n"
        "- Tổng tỷ trọng tất cả Objectives phải bằng 100%/The sum of all objectives is 100%"
    )
    ws["A4"].font = RED_BOLD
    ws["A4"].alignment = Alignment(wrap_text=True, vertical="top")
    ws.row_dimensions[4].height = 48
    ws["A5"] = "Nhân viên nhận xét tổng quan (*)\nOverall Employee Comment"
    ws["A6"] = "Nhu cầu phát triển của nhân viên (*)\nDevelopment Needs"
    for r in (5, 6):
        ws.cell(row=r, column=1).font = Font(bold=True)
        ws.cell(row=r, column=1).alignment = Alignment(wrap_text=True, vertical="top")
        ws.merge_cells(start_row=r, start_column=2, end_row=r, end_column=6)
        for c in range(1, 7):
            ws.cell(row=r, column=c).border = BORDER
        ws.row_dimensions[r].height = 52
    ws["B5"] = overall_comment
    ws["B6"] = development_needs
    for r in (5, 6):
        ws.cell(row=r, column=2).alignment = Alignment(wrap_text=True, vertical="top")

    # ===== Bang Objective / KPI =====
    HEAD_ROW = 8
    headers = [
        "Objective (*)",
        "Tỷ trọng Objective/Objective Proportion* (%)",
        "KPI (*)",
        "Tỷ trọng KPI/KPI Proportion* (%)",
        "Nhân viên tự nhận xét/Employee self-assessment*",
        "Nhân viên đánh giá/Self-KPI rating*",
    ]
    for c, h in enumerate(headers, 1):
        cell = ws.cell(row=HEAD_ROW, column=c, value=h)
        cell.fill = APPRAISAL_BLUE
        cell.font = Font(bold=True, size=10)
        cell.border = BORDER
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    ws.row_dimensions[HEAD_ROW].height = 30

    def kpi_row(r: int, kpi: models.KPI | None):
        for c in range(1, 7):
            ws.cell(row=r, column=c).border = BORDER
            ws.cell(row=r, column=c).alignment = Alignment(vertical="top", wrap_text=True)
        if kpi is None:
            return
        ws.row_dimensions[r].height = 58
        ws.cell(row=r, column=3, value=kpi.name)
        ws.cell(row=r, column=4, value=kpi.weight)
        ws.cell(row=r, column=5, value=self_assessment_comments.get(kpi.id, _progress_note(kpi)))
        # cot 6 (Self-KPI rating) de trong cho nhan vien tu cham theo thang diem cong ty

    r = HEAD_ROW + 1
    groups = [(o, [k for k in kpis if k.objective_id == o.id]) for o in objectives]
    ungrouped = [k for k in kpis if k.objective_id is None]
    if ungrouped:
        groups.append((None, ungrouped))

    if not kpis:
        # Template trong: 8 dong ke san de dien tay
        for _ in range(8):
            kpi_row(r, None)
            r += 1
    else:
        for obj, group in groups:
            if not group:
                continue
            start = r
            for k in group:
                kpi_row(r, k)
                r += 1
            end = r - 1
            if end > start:
                ws.merge_cells(start_row=start, start_column=1, end_row=end, end_column=1)
                ws.merge_cells(start_row=start, start_column=2, end_row=end, end_column=2)
            ws.cell(row=start, column=1, value=obj.name if obj else "(Chưa gắn Objective)")
            ws.cell(row=start, column=2, value=obj.weight if obj else "")
            ws.cell(row=start, column=1).font = Font(bold=True)
            ws.cell(row=start, column=1).alignment = Alignment(vertical="center", wrap_text=True)
            ws.cell(row=start, column=2).alignment = Alignment(vertical="center", horizontal="center")

    for i, w in enumerate([34, 16, 40, 14, 58, 16], 1):
        ws.column_dimensions[ws.cell(row=HEAD_ROW, column=i).column_letter].width = w

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def export_evaluation_excel(db: Session, user_id: int = 1, cycle_id: int | None = None) -> bytes:
    kpis = kpi_service.get_active_kpis(db, user_id, cycle_id=cycle_id)
    today = date.today()
    wb = Workbook()

    # ===== Sheet 1: Tong quan KPI =====
    ws = wb.active
    ws.title = "Tổng quan KPI"
    ws["A1"] = f"BÁO CÁO ĐÁNH GIÁ KPI NĂM {kpis[0].year if kpis else today.year}"
    ws["A1"].font = TITLE_FONT
    ws["A2"] = f"Ngày xuất: {today.isoformat()}"
    headers = ["STT", "Mục tiêu (Objective)", "KPI", "Mô tả / Chỉ tiêu",
               "Trọng số trong mục tiêu (%)", "Đơn vị", "Chỉ tiêu số", "Thực đạt", "Deadline",
               "Tiến độ (%)", "Kỳ vọng (%)", "Chênh lệch", "Trạng thái"]
    ws.append([])
    ws.append(headers)
    _style_header(ws, 4, len(headers))
    # nhom theo muc tieu de doc de hon
    kpis_sorted = sorted(kpis, key=lambda k: (k.objective_name or "ZZZ (chưa gắn mục tiêu)", k.id))
    for i, k in enumerate(kpis_sorted, 1):
        health, gap = kpi_service.health_of(k, today)
        exp = kpi_service.expected_progress(k, today)
        label = {"green": "Đúng tiến độ", "yellow": "Cần chú ý", "red": "Rủi ro"}[health]
        if k.progress > 100:
            label = "Vượt chỉ tiêu ★"
        ws.append([i, k.objective_name or "(chưa gắn mục tiêu)", k.name, k.target or k.description,
                   k.weight, k.unit, k.target_value, k.current_value,
                   str(k.deadline or f"{k.year}-12-31"), k.progress, exp, gap, label])
        row = ws.max_row
        for c in range(1, len(headers) + 1):
            ws.cell(row=row, column=c).border = BORDER
            ws.cell(row=row, column=c).alignment = Alignment(vertical="top", wrap_text=True)
        ws.cell(row=row, column=13).fill = (
            PatternFill("solid", fgColor="D9E1F2") if k.progress > 100 else HEALTH_FILLS[health]
        )
    widths = [5, 26, 32, 34, 13, 10, 10, 10, 12, 11, 11, 11, 14]
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[ws.cell(row=4, column=i).column_letter].width = w
    dash = kpi_service.build_dashboard(db, user_id, cycle_id=cycle_id)
    ws.append([])
    ws.append(["", "TỔNG TIẾN ĐỘ NĂM (theo trọng số mục tiêu, KPI vượt tính tối đa 100%)",
               "", "", "", "", "", "", "", dash.overall_progress])
    ws.cell(row=ws.max_row, column=2).font = Font(bold=True)
    ws.cell(row=ws.max_row, column=10).font = Font(bold=True)
    # bang tom tat theo muc tieu
    ws.append([])
    ws.append(["", "TIẾN ĐỘ THEO MỤC TIÊU (OBJECTIVE)", "Trọng số (%)", "Số KPI", "Tiến độ (%)"])
    ws.cell(row=ws.max_row, column=2).font = Font(bold=True, color="1F4E79")
    for o in dash.objectives:
        ws.append(["", o.name, o.weight, o.kpi_count, o.progress])

    # ===== Sheet 2: Bang chung cong viec =====
    ws2 = wb.create_sheet("Bằng chứng công việc")
    headers2 = ["Ngày", "Đầu việc", "Chi tiết", "Trạng thái", "KPI", "+Thực đạt (theo đơn vị KPI)", "Nguồn", "Nguồn gốc dữ liệu"]
    ws2.append(headers2)
    _style_header(ws2, 1, len(headers2))
    items = list(
        db.scalars(
            select(models.WorkItem)
            .where(models.WorkItem.user_id == user_id, models.WorkItem.confirmed == True)  # noqa: E712
            .order_by(models.WorkItem.work_date.desc().nullslast(), models.WorkItem.created_at.desc())
        )
    )
    for w in items:
        ws2.append([
            str(w.work_date or w.created_at.date()), w.title, w.detail,
            schemas.STATUS_LABELS.get(w.status, w.status),
            w.kpi.name if w.kpi else "(việc phát sinh - chưa gắn KPI)",
            w.progress_delta, w.source, w.source_ref,
        ])
        for c in range(1, len(headers2) + 1):
            ws2.cell(row=ws2.max_row, column=c).border = BORDER
            ws2.cell(row=ws2.max_row, column=c).alignment = Alignment(vertical="top", wrap_text=True)
    for i, w in enumerate([12, 35, 30, 12, 30, 12, 10, 30], 1):
        ws2.column_dimensions[ws2.cell(row=1, column=i).column_letter].width = w

    # ===== Sheet 3: Phat sinh ngoai ke hoach =====
    ws3 = wb.create_sheet("Phát sinh ngoài kế hoạch")
    headers3 = ["Ngày", "Đầu việc", "Chi tiết", "Nguồn gốc"]
    ws3.append(headers3)
    _style_header(ws3, 1, len(headers3))
    for w in items:
        if w.status == "phat_sinh":
            ws3.append([str(w.work_date or w.created_at.date()), w.title, w.detail, w.source_ref])
            for c in range(1, len(headers3) + 1):
                ws3.cell(row=ws3.max_row, column=c).border = BORDER
    for i, w in enumerate([12, 40, 40, 30], 1):
        ws3.column_dimensions[ws3.cell(row=1, column=i).column_letter].width = w

    # ===== Sheet 4: Lich su thay doi KPI =====
    ws4 = wb.create_sheet("Lịch sử thay đổi KPI")
    headers4 = ["Ngày", "KPI", "Trường thay đổi", "Giá trị cũ", "Giá trị mới", "Lý do"]
    ws4.append(headers4)
    _style_header(ws4, 1, len(headers4))
    logs = list(
        db.scalars(select(models.KPIChangeLog).order_by(models.KPIChangeLog.changed_at.desc()))
    )
    for lg in logs:
        kpi = db.get(models.KPI, lg.kpi_id)
        ws4.append([
            lg.changed_at.date().isoformat(), kpi.name if kpi else "KPI đã bị xóa hoặc không còn truy cập được",
            lg.field, lg.old_value, lg.new_value, lg.reason,
        ])
        for c in range(1, len(headers4) + 1):
            ws4.cell(row=ws4.max_row, column=c).border = BORDER
    for i, w in enumerate([12, 35, 16, 25, 25, 30], 1):
        ws4.column_dimensions[ws4.cell(row=1, column=i).column_letter].width = w

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


# ===== Self-review export =====

_SECTION_FILL = PatternFill("solid", fgColor="4472C4")
_SECTION_FONT = Font(bold=True, color="FFFFFF", size=11)


def export_self_review_excel(period_label: str, content: str) -> bytes:
    """Xuat ban tu danh gia thanh file Excel co dinh dang dep."""
    wb = Workbook()
    ws = wb.active
    ws.title = "Tự đánh giá"
    ws.column_dimensions["A"].width = 3
    ws.column_dimensions["B"].width = 90

    # Tieu de chinh
    ws.merge_cells("A1:B1")
    c = ws["A1"]
    c.value = f"BẢN TỰ ĐÁNH GIÁ — {period_label.upper()}"
    c.font = Font(bold=True, size=14, color="FFFFFF")
    c.fill = PatternFill("solid", fgColor="1F4E79")
    c.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 30

    ws.merge_cells("A2:B2")
    c = ws["A2"]
    c.value = f"Ngày xuất: {date.today().isoformat()}"
    c.font = Font(italic=True, color="808080", size=10)
    c.alignment = Alignment(horizontal="right")

    row = 4
    pending: list[str] = []

    def _flush(r: int) -> int:
        """Ghi cac dong noi dung dang cho vao sheet, tra ve row moi."""
        for ln in pending:
            stripped = ln.strip()
            if not stripped:
                continue
            # Strip markdown bullet/bold markers
            stripped = stripped.lstrip("- *")
            stripped = stripped.replace("**", "").replace("*", "").replace("`", "")
            ws[f"B{r}"] = stripped
            ws[f"B{r}"].font = Font(size=10)
            ws[f"B{r}"].alignment = Alignment(wrap_text=True, vertical="top")
            ws.row_dimensions[r].height = 15
            r += 1
        pending.clear()
        return r

    for line in content.split("\n"):
        if line.startswith("## ") or line.startswith("### "):
            row = _flush(row)
            row += 1  # khoang cach truoc section
            header_text = line.lstrip("#").strip()
            ws.merge_cells(f"A{row}:B{row}")
            c = ws[f"A{row}"]
            c.value = header_text
            c.font = _SECTION_FONT
            c.fill = _SECTION_FILL
            c.alignment = Alignment(vertical="center")
            ws.row_dimensions[row].height = 22
            row += 1
        else:
            pending.append(line)

    _flush(row)

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def _md_strip(text: str) -> str:
    """Bỏ markdown inline (bold/italic/code/link) trả về plain text."""
    import re
    text = re.sub(r"\*\*(.+?)\*\*", r"\1", text)  # **bold**
    text = re.sub(r"\*(.+?)\*", r"\1", text)       # *italic*
    text = re.sub(r"`(.+?)`", r"\1", text)          # `code`
    text = re.sub(r"\[(.+?)\]\(.+?\)", r"\1", text) # [link](url)
    return text


def _export_report_pdf_legacy(period_label: str, content: str) -> bytes:
    """Xuat noi dung bao cao (markdown) ra PDF co font Unicode."""
    import re
    from .export_service import _find_pdf_font
    from fpdf import FPDF

    reg, bold = _find_pdf_font()
    pdf = FPDF()
    pdf.set_auto_page_break(True, margin=15)
    if reg:
        pdf.add_font("Main", "", reg)
        pdf.add_font("Main", "B", bold or reg)
        fam = "Main"
    else:
        fam = "Helvetica"
    pdf.add_page()
    epw = pdf.w - 2 * pdf.l_margin

    def _safe_cell(text: str, font_style: str = "", font_size: int = 10,
                   line_h: float = 5, color: tuple = (0, 0, 0)):
        """Render một đoạn text, bỏ qua dòng nếu gặp lỗi font."""
        try:
            pdf.set_font(fam, font_style, font_size)
            pdf.set_text_color(*color)
            # Lọc ký tự ngoài BMP (emoji, surrogate) để tránh lỗi fpdf2
            safe = re.sub(r"[^ -￿]", "?", text)
            pdf.multi_cell(epw, line_h, safe, new_x="LMARGIN", new_y="NEXT")
            pdf.set_text_color(0, 0, 0)
        except Exception:
            pass  # bỏ qua dòng lỗi, không làm hỏng cả file

    # Tieu de chinh
    _safe_cell(f"BẢN TỰ ĐÁNH GIÁ — {period_label.upper()}", "B", 15, 9, (31, 78, 121))
    try:
        pdf.set_font(fam, "", 9)
        pdf.set_text_color(130, 130, 130)
        pdf.cell(0, 6, f"Ngày xuất: {date.today().isoformat()}", new_x="LMARGIN", new_y="NEXT")
        pdf.set_text_color(0, 0, 0)
    except Exception:
        pass
    pdf.ln(3)

    for line in content.split("\n"):
        if line.startswith("## "):
            pdf.ln(3)
            _safe_cell(_md_strip(line[3:].strip()), "B", 12, 8, (31, 78, 121))
            pdf.ln(1)
        elif line.startswith("### "):
            pdf.ln(1)
            _safe_cell(_md_strip(line[4:].strip()), "B", 10, 6, (68, 114, 196))
        elif line.startswith("- ") or line.startswith("* "):
            _safe_cell("  • " + _md_strip(line[2:].strip()), "", 10, 5)
        elif line.strip():
            _safe_cell(_md_strip(line.strip()), "", 10, 5)
        else:
            pdf.ln(2)

    out = pdf.output()
    return bytes(out)


def _report_pdf_plain(text: str) -> str:
    """Return readable plain text from common Markdown/HTML inline markers."""
    import re

    text = re.sub(r"!\[(.*?)\]\(.+?\)", r"\1", text)
    text = re.sub(r"\[(.+?)\]\(.+?\)", r"\1", text)
    text = re.sub(r"<br\s*/?>", "\n", text, flags=re.IGNORECASE)
    text = re.sub(r"</?u>", "", text, flags=re.IGNORECASE)
    text = re.sub(r"<[^>]+>", "", text)
    text = re.sub(r"\*\*(.+?)\*\*", r"\1", text)
    text = re.sub(r"__(.+?)__", r"\1", text)
    text = re.sub(r"\*(.+?)\*", r"\1", text)
    text = re.sub(r"_(.+?)_", r"\1", text)
    text = re.sub(r"`(.+?)`", r"\1", text)
    return text.strip()


def _report_pdf_markdown_text(text: str) -> str:
    """Keep supported Markdown emphasis, but remove links/tags unsafe for PDF."""
    import re

    text = re.sub(r"!\[(.*?)\]\(.+?\)", r"\1", text)
    text = re.sub(r"\[(.+?)\]\(.+?\)", r"\1", text)
    text = re.sub(r"<br\s*/?>", "\n", text, flags=re.IGNORECASE)
    text = re.sub(r"</?u>", "", text, flags=re.IGNORECASE)
    text = re.sub(r"<[^>]+>", "", text)
    text = re.sub(r"`(.+?)`", r"\1", text)
    text = re.sub(r"(?<!\*)\*(?!\*)(.+?)(?<!\*)\*(?!\*)", r"\1", text)
    text = re.sub(r"(?<!_)_(?!_)(.+?)(?<!_)_(?!_)", r"\1", text)
    return text.strip()


def _report_pdf_safe(text: str) -> str:
    """Remove non-BMP glyphs that can break some PDF font backends."""
    import re

    return re.sub(r"[\U00010000-\U0010ffff]", "", str(text or "")).strip()


def _report_pdf_split_table_row(line: str) -> list[str]:
    return [cell.strip() for cell in line.strip().strip("|").split("|")]


def _report_pdf_is_table_separator(line: str) -> bool:
    cells = _report_pdf_split_table_row(line)
    return bool(cells) and all(cell and set(cell.replace(":", "").strip()) <= {"-"} for cell in cells)


def export_report_pdf(period_label: str, content: str, period_type: str = "report") -> bytes:
    """Export saved report Markdown to a polished PDF using the current report body."""
    import re
    from .export_service import _find_pdf_font
    from fpdf import FPDF

    class ReportPDF(FPDF):
        def footer(self):
            self.set_y(-12)
            self.set_font(fam, "", 8)
            self.set_text_color(130, 130, 130)
            self.cell(0, 7, str(self.page_no()), align="C")

    reg, bold = _find_pdf_font()
    pdf = ReportPDF()
    pdf.set_auto_page_break(True, margin=15)
    if reg:
        pdf.add_font("Main", "", reg)
        pdf.add_font("Main", "B", bold or reg)
        pdf.add_font("Main", "I", reg)
        pdf.add_font("Main", "BI", bold or reg)
        fam = "Main"
    else:
        fam = "Helvetica"
    pdf.add_page()
    epw = pdf.w - 2 * pdf.l_margin

    def _safe_multi(
        text: str,
        style: str = "",
        size: float = 10,
        line_h: float = 5.5,
        color: tuple[int, int, int] = (15, 23, 42),
        indent: float = 0,
        markdown: bool = True,
        fill: tuple[int, int, int] | None = None,
    ):
        safe = _report_pdf_safe(text)
        if not safe:
            return
        try:
            pdf.set_font(fam, style, size)
            pdf.set_text_color(*color)
            if fill:
                pdf.set_fill_color(*fill)
            pdf.set_x(pdf.l_margin + indent)
            pdf.multi_cell(
                epw - indent,
                line_h,
                safe,
                fill=bool(fill),
                markdown=markdown,
                new_x="LMARGIN",
                new_y="NEXT",
            )
        except Exception:
            try:
                pdf.set_font(fam, style if style in {"", "B"} else "", size)
                pdf.set_text_color(*color)
                pdf.set_x(pdf.l_margin + indent)
                pdf.multi_cell(epw - indent, line_h, _report_pdf_plain(safe), new_x="LMARGIN", new_y="NEXT")
            except Exception:
                pass
        finally:
            pdf.set_text_color(0, 0, 0)

    def _render_table(table_lines: list[str]):
        rows = [
            _report_pdf_split_table_row(line)
            for line in table_lines
            if not _report_pdf_is_table_separator(line)
        ]
        rows = [[_report_pdf_plain(cell) for cell in row] for row in rows if any(cell.strip() for cell in row)]
        if not rows:
            return
        n_cols = max(len(row) for row in rows)
        rows = [row + [""] * (n_cols - len(row)) for row in rows]
        try:
            pdf.ln(1)
            pdf.set_font(fam, "", 8)
            with pdf.table(
                width=epw,
                col_widths=[1] * n_cols,
                first_row_as_headings=True,
                line_height=5,
                text_align="LEFT",
            ) as table:
                for row_values in rows:
                    row = table.row()
                    for value in row_values:
                        row.cell(_report_pdf_safe(value))
            pdf.ln(2)
        except Exception:
            for row_values in rows:
                _safe_multi(" | ".join(row_values), size=8.5, line_h=4.8, markdown=False)

    report_kind = "Bản tự đánh giá KPI" if period_type == "self_review" else "Báo cáo KPI"
    _safe_multi(report_kind.upper(), "B", 16, 9, (255, 255, 255), markdown=False, fill=(31, 78, 121))
    _safe_multi(period_label, "B", 13, 7, (31, 78, 121), markdown=False)
    _safe_multi(f"Ngày xuất: {date.today().isoformat()} · Định dạng: PDF", "", 9, 5, (100, 116, 139), markdown=False)
    pdf.ln(3)

    lines = content.replace("\r\n", "\n").replace("\r", "\n").split("\n")
    i = 0
    while i < len(lines):
        line = lines[i].rstrip()
        stripped = line.strip()
        if not stripped:
            pdf.ln(2)
            i += 1
            continue

        if stripped.startswith("|") and i + 1 < len(lines) and _report_pdf_is_table_separator(lines[i + 1]):
            table_lines = []
            while i < len(lines) and lines[i].strip().startswith("|"):
                table_lines.append(lines[i])
                i += 1
            _render_table(table_lines)
            continue

        if stripped.startswith("# "):
            _safe_multi(_report_pdf_plain(stripped[2:]), "B", 14, 8, (31, 78, 121), markdown=False)
            pdf.ln(1)
        elif stripped.startswith("## "):
            pdf.ln(2)
            _safe_multi(_report_pdf_plain(stripped[3:]), "B", 12, 7, (31, 78, 121), markdown=False, fill=(232, 240, 254))
            pdf.ln(1)
        elif stripped.startswith("### "):
            pdf.ln(1)
            _safe_multi(_report_pdf_plain(stripped[4:]), "B", 10.5, 6, (68, 114, 196), markdown=False)
        elif stripped.startswith(">"):
            quote = stripped.lstrip(">").strip()
            _safe_multi(_report_pdf_markdown_text(quote), "I", 9.5, 5.3, (71, 85, 105), indent=4)
        elif re.match(r"^[-*]\s+", stripped):
            body = re.sub(r"^[-*]\s+", "", stripped)
            _safe_multi("• " + _report_pdf_markdown_text(body), "", 10, 5.3, indent=4)
        elif re.match(r"^\d+\.\s+", stripped):
            _safe_multi(_report_pdf_markdown_text(stripped), "", 10, 5.3, indent=4)
        elif set(stripped) <= {"-", "_", "*"} and len(stripped) >= 3:
            pdf.ln(1)
            pdf.set_draw_color(203, 213, 225)
            pdf.line(pdf.l_margin, pdf.get_y(), pdf.w - pdf.r_margin, pdf.get_y())
            pdf.ln(3)
        else:
            paragraph = [stripped]
            i += 1
            while i < len(lines):
                nxt = lines[i].strip()
                if (
                    not nxt
                    or nxt.startswith(("#", ">", "|"))
                    or re.match(r"^[-*]\s+", nxt)
                    or re.match(r"^\d+\.\s+", nxt)
                ):
                    break
                paragraph.append(nxt)
                i += 1
            _safe_multi(_report_pdf_markdown_text(" ".join(paragraph)), "", 10, 5.4)
            continue
        i += 1

    out = pdf.output()
    return bytes(out)
