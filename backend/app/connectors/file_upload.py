"""Doc file Excel/CSV nguoi dung upload (timesheet, log cong viec, danh sach KPI)."""
import csv
import io
from datetime import date, datetime

from openpyxl import load_workbook


def _rows_from_bytes(filename: str, content: bytes) -> list[list[str]]:
    if filename.lower().endswith((".xlsx", ".xlsm")):
        wb = load_workbook(io.BytesIO(content), read_only=True, data_only=True)
        rows: list[list[str]] = []
        for ws in wb.worksheets:
            # Some exports (for example ManageEngine reports) incorrectly set
            # dimension="A1"; read_only mode then sees a blank sheet unless reset.
            if hasattr(ws, "reset_dimensions"):
                ws.reset_dimensions()
            sheet_rows = [
                ["" if c is None else (c.isoformat() if isinstance(c, (date, datetime)) else str(c)) for c in row]
                for row in ws.iter_rows(values_only=True)
            ]
            if any(any(str(c).strip() for c in row) for row in sheet_rows):
                if len(wb.worksheets) > 1:
                    rows.append([f"Sheet: {ws.title}"])
                rows.extend(sheet_rows)
        return rows
    # CSV: thu utf-8-sig truoc (Excel VN hay xuat BOM)
    for enc in ("utf-8-sig", "utf-8", "cp1258", "latin-1"):
        try:
            text = content.decode(enc)
            break
        except UnicodeDecodeError:
            continue
    return [row for row in csv.reader(io.StringIO(text))]


def parse_worklog_file(filename: str, content: bytes) -> list[dict]:
    """Chuyen file log cong viec thanh danh sach hoat dong chuan hoa cho Agent phan loai."""
    rows = _rows_from_bytes(filename, content)
    if not rows:
        return []
    header = [h.strip().lower() for h in rows[0]]

    def col(*names: str) -> int | None:
        for n in names:
            for i, h in enumerate(header):
                if n in h:
                    return i
        return None

    i_date = col("ngày", "ngay", "date")
    i_task = col("công việc", "cong viec", "task", "việc", "nội dung", "noi dung")
    i_status = col("trạng thái", "trang thai", "status")
    i_note = col("ghi chú", "ghi chu", "note")

    out = []
    for line_no, row in enumerate(rows[1:], start=2):
        if not any(str(c).strip() for c in row):
            continue
        task = row[i_task] if i_task is not None and i_task < len(row) else " | ".join(map(str, row))
        if not str(task).strip():
            continue
        d = ""
        if i_date is not None and i_date < len(row):
            d = str(row[i_date])[:10]
        status = row[i_status] if i_status is not None and i_status < len(row) else ""
        note = row[i_note] if i_note is not None and i_note < len(row) else ""
        text = f"{task}"
        if status:
            text += f" — trạng thái: {status}"
        if note:
            text += f" — {note}"
        out.append(
            {
                "source": "csv",
                "date": d or datetime.now().date().isoformat(),
                "text": text,
                "ref": f"file {filename}, dòng {line_no}",
            }
        )
    return out


def parse_appraisal_file(filename: str, content: bytes) -> dict | None:
    """Nhan dien va doc file theo MAU PERFORMANCE APPRAISAL cua cong ty.

    Cau truc mau: dong tieu de "Objective (*) | Ty trong Objective (%) | KPI (*) |
    Ty trong KPI (%) | Nhan vien tu nhan xet | Self-KPI rating".
    O Objective bi merge -> cac dong sau de trong, tu mang gia tri dong tren xuong.

    Tra ve {"objectives": [{"name", "weight", "kpis": [{"name", "weight", "note"}]}]}
    hoac None neu file KHONG theo mau nay (de fallback sang parse_kpi_file).
    """
    rows = _rows_from_bytes(filename, content)
    header_idx = None
    for i, row in enumerate(rows[:30]):
        cells = [str(c).strip().lower() for c in row if c and str(c).strip()]
        # Header that: "objective" va "kpi" o cell RIENG BIET (khong phai cung mot dong instruction dai)
        # Dieu kien: co it nhat 1 cell chi chua "objective" (khong co "kpi") va 1 cell chi chua "kpi" (khong co "objective")
        has_obj_cell = any("objective" in c and "kpi" not in c for c in cells)
        has_kpi_cell = any("kpi" in c and "objective" not in c for c in cells)
        has_weight = any("tỷ trọng" in c or "ty trong" in c or "proportion" in c for c in cells)
        if has_obj_cell and has_kpi_cell and has_weight:
            header_idx = i
            break
    if header_idx is None:
        return None

    header = [str(h).strip().lower() for h in rows[header_idx]]

    def find(pred) -> int | None:
        for i, h in enumerate(header):
            if h and pred(h):
                return i
        return None

    is_w = lambda h: "proportion" in h or "tỷ trọng" in h or "ty trong" in h  # noqa: E731
    i_objw = find(lambda h: "objective" in h and is_w(h))
    i_obj = find(lambda h: "objective" in h and not is_w(h))
    i_kpiw = find(lambda h: "kpi" in h and is_w(h))
    i_kpi = find(
        lambda h: "kpi" in h and not is_w(h) and "rating" not in h
        and "nhận xét" not in h and "assessment" not in h and "đánh giá" not in h
    )
    i_note = find(lambda h: "assessment" in h or "nhận xét" in h)
    if i_obj is None or i_kpi is None:
        return None

    def cell(row: list, idx: int | None) -> str:
        return str(row[idx]).strip() if idx is not None and idx < len(row) and row[idx] else ""

    def to_num(raw: str) -> float | None:
        raw = raw.replace("%", "").replace(",", ".").strip()
        try:
            return float(raw)
        except ValueError:
            return None

    objectives: list[dict] = []
    by_name: dict[str, dict] = {}
    current = None  # objective dang mang xuong (xu ly o merge)
    for row in rows[header_idx + 1:]:
        obj_name = cell(row, i_obj)
        kpi_name = cell(row, i_kpi)
        if obj_name:
            key = obj_name.lower()
            if key not in by_name:
                by_name[key] = {"name": obj_name, "weight": to_num(cell(row, i_objw)) or 0.0, "kpis": []}
                objectives.append(by_name[key])
            elif to_num(cell(row, i_objw)) is not None:
                by_name[key]["weight"] = to_num(cell(row, i_objw))
            current = by_name[key]
        if not kpi_name:
            continue
        if current is None:  # KPI truoc khi co Objective nao -> file sai cau truc
            return None
        current["kpis"].append(
            {
                "name": kpi_name,
                "weight": to_num(cell(row, i_kpiw)) or 0.0,
                "note": cell(row, i_note),
            }
        )
    if not objectives or not any(o["kpis"] for o in objectives):
        return None
    return {"objectives": objectives}


def parse_kpi_file(filename: str, content: bytes) -> list[dict]:
    """Doc file danh sach KPI: ten | mo ta | chi tieu | trong so | deadline."""
    rows = _rows_from_bytes(filename, content)
    if not rows:
        return []
    header = [h.strip().lower() for h in rows[0]]

    def col(*names: str) -> int | None:
        for n in names:
            for i, h in enumerate(header):
                if n in h:
                    return i
        return None

    i_name = col("tên", "ten", "kpi", "name")
    i_desc = col("mô tả", "mo ta", "description")
    i_target = col("chỉ tiêu", "chi tieu", "target", "đo lường", "do luong")
    i_weight = col("trọng số", "trong so", "weight")
    i_deadline = col("deadline", "hạn", "han")

    def cell(row: list, idx: int | None) -> str:
        return str(row[idx]).strip() if idx is not None and idx < len(row) and row[idx] else ""

    out = []
    for row in rows[1:]:
        name = cell(row, i_name) or (str(row[0]).strip() if row else "")
        if not name:
            continue
        weight = 0.0
        w_raw = cell(row, i_weight).replace("%", "").replace(",", ".")
        try:
            weight = float(w_raw)
        except ValueError:
            pass
        deadline = None
        d_raw = cell(row, i_deadline)[:10]
        try:
            deadline = date.fromisoformat(d_raw) if d_raw else None
        except ValueError:
            pass
        out.append(
            {
                "name": name,
                "description": cell(row, i_desc),
                "target": cell(row, i_target),
                "weight": weight,
                "deadline": deadline,
            }
        )
    return out
